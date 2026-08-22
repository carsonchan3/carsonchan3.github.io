from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = PROJECT_ROOT / "client" / "src" / "pages" / "Equipment.tsx"
OUTPUT_FILE = Path("/home/ubuntu/VLI_Product_Catalogue_Template.xlsx")


def product_records():
    source = SOURCE_FILE.read_text(encoding="utf-8")
    catalogue = source.split("export const catalogueItems = [", 1)[1].split("];\n\nexport default", 1)[0]
    object_blocks = re.findall(r"\{\n(.*?)\n  \},", catalogue, flags=re.DOTALL)
    fields = ["number", "sourceId", "model", "name", "category", "description", "price", "image", "imageAlt"]
    records = []
    for block in object_blocks:
        item = {}
        for field in fields:
            match = re.search(rf'{field}: "([^"]*)"', block)
            if not match:
                raise ValueError(f"Missing {field} in a catalogue record")
            item[field] = match.group(1)
        item["price_hkd"] = int(re.sub(r"[^0-9]", "", item["price"]))
        records.append(item)
    if len(records) != 22:
        raise ValueError(f"Expected 22 Product catalogue items, found {len(records)}")
    return records


def apply_sheet_style(sheet, widths):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_workbook(records):
    workbook = Workbook()
    guide = workbook.active
    guide.title = "Read Me"
    catalog = workbook.create_sheet("Product Catalog")
    mapping = workbook.create_sheet("Field Guide")

    dark = "1C1D20"
    charcoal = "27282B"
    turquoise = "40E0D0"
    white = "FFFFFF"
    muted = "D9E1E5"
    border = Border(
        left=Side(style="thin", color="465057"),
        right=Side(style="thin", color="465057"),
        top=Side(style="thin", color="465057"),
        bottom=Side(style="thin", color="465057"),
    )

    guide.merge_cells("A1:H1")
    guide["A1"] = "Velocity Lab Innovation — Product Catalogue Template"
    guide["A1"].fill = PatternFill("solid", fgColor=dark)
    guide["A1"].font = Font(color=white, bold=True, size=16)
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 34
    guide["A3"] = "Purpose"
    guide["B3"] = "Use the Product Catalog sheet to maintain the content that appears on VLI’s Product page. It is prefilled with the 22 items currently published."
    guide["A4"] = "Safe editing"
    guide["B4"] = "Do not change Source ID for existing products. It is the stable product reference used for catalogue and quote-cart records."
    guide["A5"] = "Prices"
    guide["B5"] = "Enter the listed price as a whole number in HKD. The site formats it as HK$ with thousands separators."
    guide["A6"] = "Images"
    guide["B6"] = "Provide a hosted image URL or managed storage path in Image URL. Use concise, descriptive Image Alt Text for accessibility."
    guide["A7"] = "Status"
    guide["B7"] = "Set Active to display the item or Hidden to retain it in the workbook without publishing it in the next catalogue update."
    guide["A8"] = "Publishing"
    guide["B8"] = "Save your completed workbook and provide it for the next site update. Editing this file alone does not change the live website automatically."
    for row in range(3, 9):
        guide[f"A{row}"].font = Font(bold=True, color=turquoise)
        guide[f"A{row}"].fill = PatternFill("solid", fgColor=charcoal)
        guide[f"A{row}"].border = border
        guide[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        guide[f"B{row}"].border = border
        guide[f"B{row}"].fill = PatternFill("solid", fgColor="F7FAFC")
        guide.row_dimensions[row].height = 42
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 105
    for column in "CDEFGH":
        guide.column_dimensions[column].width = 3
    guide.sheet_view.showGridLines = False

    headers = [
        "Source ID (do not change)",
        "Product No.",
        "Model",
        "Product Name",
        "Category",
        "Description",
        "Price (HKD)",
        "Image URL",
        "Image Alt Text",
        "Status",
        "Internal Notes",
    ]
    catalog.append(headers)
    for cell in catalog[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    catalog.row_dimensions[1].height = 34
    for item in records:
        catalog.append([
            item["sourceId"],
            int(item["number"]),
            item["model"],
            item["name"],
            item["category"],
            item["description"],
            item["price_hkd"],
            item["image"],
            item["imageAlt"],
            "Active",
            "",
        ])
    for row in catalog.iter_rows(min_row=2, max_row=catalog.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        row[6].number_format = 'HK$#,##0'
        catalog.row_dimensions[row[0].row].height = 42
    catalog.auto_filter.ref = f"A1:K{catalog.max_row}"
    catalog.freeze_panes = "A2"
    catalog.sheet_view.showGridLines = False
    for column, width in {
        "A": 22, "B": 13, "C": 22, "D": 31, "E": 22, "F": 55, "G": 16, "H": 57, "I": 47, "J": 13, "K": 32,
    }.items():
        catalog.column_dimensions[column].width = width
    status_validation = DataValidation(type="list", formula1='"Active,Hidden"', allow_blank=False)
    catalog.add_data_validation(status_validation)
    status_validation.add(f"J2:J{catalog.max_row + 200}")
    catalog.conditional_formatting.add(
        f"J2:J{catalog.max_row + 200}",
        FormulaRule(formula=["$J2=\"Hidden\""], fill=PatternFill("solid", fgColor=muted)),
    )

    mapping_headers = ["Workbook column", "Website field", "Required", "Guidance"]
    mapping.append(mapping_headers)
    mapping_rows = [
        ["Source ID (do not change)", "sourceId", "Yes", "A stable reference used by the Product page and quote cart. Keep unique and unchanged for existing items."],
        ["Product No.", "number", "Yes", "The visible product reference shown as #number."],
        ["Model", "model", "Yes", "Manufacturer or internal model identifier."],
        ["Product Name", "name", "Yes", "Short public-facing product title."],
        ["Category", "category", "Yes", "Public grouping such as Drone platform, Drone power, Charging equipment, or Competition venue."],
        ["Description", "description", "Yes", "Brief public-facing description. Aim for one concise sentence."],
        ["Price (HKD)", "price", "Yes", "Whole-number listed price. The site renders this as HK$X,XXX."],
        ["Image URL", "image", "Yes", "Hosted URL or managed storage path for the product image."],
        ["Image Alt Text", "imageAlt", "Yes", "Describe the visible product image for assistive technologies."],
        ["Status", "publication flag", "Yes", "Active = include in next catalogue update; Hidden = retain but do not publish."],
        ["Internal Notes", "not published", "No", "Use for sourcing, stock, supplier, or review notes. It is not displayed publicly."],
    ]
    for row in mapping_rows:
        mapping.append(row)
    for cell in mapping[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color=white, bold=True)
        cell.border = border
    for row in mapping.iter_rows(min_row=2, max_row=mapping.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        mapping.row_dimensions[row[0].row].height = 42
    mapping.sheet_view.showGridLines = False
    mapping.freeze_panes = "A2"
    mapping.auto_filter.ref = mapping.dimensions
    for column, width in {"A": 26, "B": 22, "C": 12, "D": 92}.items():
        mapping.column_dimensions[column].width = width

    workbook.active = 0
    workbook.save(OUTPUT_FILE)


if __name__ == "__main__":
    records = product_records()
    build_workbook(records)
    print(f"Created {OUTPUT_FILE} with {len(records)} current catalogue items.")
